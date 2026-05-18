import logging
import pandas as pd
import json
import os
import glob
import openpyxl
import yaml
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class CalcularEconomicidadeProcessor:
    def __init__(self, session=None):
        self.logger = logging.getLogger(__name__)
        self.unidades_path = "cache/unidades_cnes.json"
        self.config_path = "config.yaml"
        
        self.unidades = self._carregar_unidades()
        self.config = self._carregar_configuracoes()

    def _carregar_unidades(self):
        if os.path.exists(self.unidades_path):
            try:
                with open(self.unidades_path, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception as e:
                self.logger.error(f"Erro ao carregar cache de unidades: {e}")
        return {}

    def _carregar_configuracoes(self):
        """
        Carrega as variáveis financeiras a partir do arquivo config.yaml estruturado.
        """
        valores_padrao = {
            "salvador_base": 11.80,
            "multiplicador_ida_volta": 2.0,
            "interior_km": 0.10,
            "interior_fixo": 8.56,
            "interior_co2": 0.192
        }
        
        if not os.path.exists(self.config_path):
            self.logger.warning(f"Arquivo '{self.config_path}' não localizado. Utilizando valores padrão de segurança.")
            return valores_padrao
            
        try:
            with open(self.config_path, 'r', encoding='utf-8') as f:
                dados_yaml = yaml.safe_load(f)
                ref = dados_yaml.get("valores_referencia", {})
                
                return {
                    "salvador_base": float(ref.get("salvador", {}).get("custo_base_passagem", 11.80)),
                    "multiplicador_ida_volta": float(ref.get("multiplicador_ida_volta", 2.0)),
                    "interior_km": float(ref.get("interior", {}).get("custo_por_km", 0.10)),
                    "interior_fixo": float(ref.get("interior", {}).get("taxa_fixa_atendimento", 8.40)),
                    "interior_co2": float(ref.get("interior", {}).get("fator_co2_por_km", 0.192))
                }
        except Exception as e:
            self.logger.error(f"Erro de leitura no config.yaml: {e}. Usando valores padrão.")
            return valores_padrao

    def encontrar_e_ler_arquivo(self) -> pd.DataFrame:
        pasta_busca = "output"
        prefixo = "teleconsultorias_ibge*"
        
        padrao = os.path.join(pasta_busca, prefixo)
        arquivos_encontrados = glob.glob(padrao)

        if not arquivos_encontrados:
            raise FileNotFoundError(f"Nenhum arquivo com o prefixo '{prefixo}' foi encontrado na pasta '{pasta_busca}/'.")

        arquivo_alvo = max(arquivos_encontrados, key=os.path.getmtime)
        self.logger.info(f"Arquivo localizado para processamento: {arquivo_alvo}")

        extensao = os.path.splitext(arquivo_alvo)[1].lower()
        if extensao in ['.xlsx', '.xls']:
            return pd.read_excel(arquivo_alvo)
        elif extensao == '.csv':
            return pd.read_csv(arquivo_alvo, sep='\t', encoding='utf-8')
        else:
            raise ValueError(f"Formato de arquivo {extensao} não suportado. Utilize Excel ou CSV.")

    def calcular_economicidade(self, df: pd.DataFrame) -> pd.DataFrame:
        self.logger.info("Iniciando o cálculo de economicidade e impacto ambiental global (Incluindo Salvador)...")
        df = df.copy()

        if 'ORIGEM_SALVADOR' in df.columns:
            df['ORIGEM_SALVADOR'] = df['ORIGEM_SALVADOR'].astype(str).str.strip().str.upper()
        else:
            raise KeyError("A coluna essencial 'ORIGEM_SALVADOR' não foi encontrada no arquivo.")

        if 'PESO_EVITACAO' in df.columns:
            df['PESO_EVITACAO'] = pd.to_numeric(df['PESO_EVITACAO'], errors='coerce').fillna(0)
        else:
            df['PESO_EVITACAO'] = 1.0

        if 'DISTANCIA_KM_OSRM' in df.columns:
            if df['DISTANCIA_KM_OSRM'].dtype == object:
                df['DISTANCIA_KM_OSRM'] = df['DISTANCIA_KM_OSRM'].astype(str).str.replace(',', '.')
            df['DISTANCIA_KM_OSRM'] = pd.to_numeric(df['DISTANCIA_KM_OSRM'], errors='coerce').fillna(0)
        else:
            df['DISTANCIA_KM_OSRM'] = 0.0

        # Inicializa colunas de resultados
        df['DISTANCIA_TOTAL_EVITADA_KM'] = 0.0
        df['CUSTO_EVITADO_RS'] = 0.0
        df['CO2_EVITADO_KG'] = 0.0
        df['CO2_EVITADO_TON'] = 0.0

        is_salvador = (df['ORIGEM_SALVADOR'] == 'SIM')
        is_interior = (df['ORIGEM_SALVADOR'] == 'NÃO') | (df['ORIGEM_SALVADOR'] == 'NAO')
        
        sal_base = self.config["salvador_base"]
        sal_mult = self.config["multiplicador_ida_volta"]
        int_km = self.config["interior_km"]
        int_fixo = self.config["interior_fixo"]
        int_co2 = self.config["interior_co2"]

        df['DISTANCIA_TOTAL_EVITADA_KM'] = df['DISTANCIA_KM_OSRM'] * sal_mult

        df.loc[is_salvador, 'CUSTO_EVITADO_RS'] = (sal_base * sal_mult) * df.loc[is_salvador, 'PESO_EVITACAO']
        
        df.loc[is_interior, 'CUSTO_EVITADO_RS'] = (df.loc[is_interior, 'DISTANCIA_TOTAL_EVITADA_KM'] * int_km) + int_fixo        

        df['CO2_EVITADO_KG'] = df['DISTANCIA_TOTAL_EVITADA_KM'] * int_co2
        df['CO2_EVITADO_TON'] = df['CO2_EVITADO_KG'] / 1000.0

        self.logger.info("Cálculos de economicidade e impacto ambiental concluídos.")
        return df

    def criar_aba_resumo(self, writer_path: str, df: pd.DataFrame):        
        wb = openpyxl.load_workbook(writer_path)
        ws = wb.create_sheet(title="Resumo de Resultados", index=0)
        ws.views.sheetView[0].showGridLines = True

        font_family = "Segoe UI"
        title_font = Font(name=font_family, size=16, bold=True, color="1F4E5B")
        section_font = Font(name=font_family, size=12, bold=True, color="2E75B6")
        header_font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E5B", end_color="1F4E5B", fill_type="solid")
        bold_font = Font(name=font_family, size=10, bold=True)
        regular_font = Font(name=font_family, size=10)
        italic_font = Font(name=font_family, size=9, italic=True)

        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
        )
        total_border = Border(top=Side(style='thin', color='1F4E5B'), bottom=Side(style='double', color='1F4E5B'))

        ws["A1"] = "Relatório de Economicidade e Impacto - Telessaúde"
        ws["A1"].font = title_font
        ws.row_dimensions[1].height = 30

        ws["A3"] = "1. Distribuição de Solicitações, Distâncias e Impacto Ambiental"
        ws["A3"].font = section_font

        headers = ["Macro Região", "Modalidade", "Qtd Solicitações", "Km Evitados (Ida/Volta)", "Custos Evitados (R$)", "CO₂ Evitado (Toneladas)"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=h)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[4].height = 28

        df['MODALIDADE'] = df['MODALIDADE'].fillna('Não Informado').astype(str)
        
        agrupado = df.groupby(['ORIGEM_SALVADOR', 'MODALIDADE']).agg(
            qtd=('ID', 'count'),
            km=('DISTANCIA_TOTAL_EVITADA_KM', 'sum'),
            custo=('CUSTO_EVITADO_RS', 'sum'),
            co2_ton=('CO2_EVITADO_TON', 'sum')
        ).reset_index()

        dados_resumo = []
        for _, reg in agrupado.iterrows():
            regiao_nome = "Salvador" if reg['ORIGEM_SALVADOR'] == 'SIM' else "Demais Regiões (Interior)"
            dados_resumo.append([regiao_nome, reg['MODALIDADE'], reg['qtd'], reg['km'], reg['custo'], reg['co2_ton']])

        dados_resumo.sort(key=lambda x: x[0], reverse=True)

        row_start = 5
        for idx, r_data in enumerate(dados_resumo):
            current_row = row_start + idx
            ws.cell(row=current_row, column=1, value=r_data[0]).alignment = Alignment(horizontal="left")
            ws.cell(row=current_row, column=2, value=r_data[1]).alignment = Alignment(horizontal="left")
            
            c_qtd = ws.cell(row=current_row, column=3, value=r_data[2])
            c_qtd.number_format = "#,##0"; c_qtd.alignment = Alignment(horizontal="right")

            c_km = ws.cell(row=current_row, column=4, value=r_data[3])
            c_km.number_format = "#,##0.00"; c_km.alignment = Alignment(horizontal="right")

            c_cost = ws.cell(row=current_row, column=5, value=r_data[4])
            c_cost.number_format = "R$ #,##0.00"; c_cost.alignment = Alignment(horizontal="right")

            c_co2 = ws.cell(row=current_row, column=6, value=r_data[5])
            c_co2.number_format = "#,##0.00"; c_co2.alignment = Alignment(horizontal="right")

            for col in range(1, 7):
                cell = ws.cell(row=current_row, column=col)
                cell.font = regular_font; cell.border = thin_border
                if idx % 2 == 1:
                    cell.fill = PatternFill(start_color="F9FBFB", end_color="F9FBFB", fill_type="solid")

        ws["A10"] = "Subtotal Salvador"
        ws["C10"] = "=SUM(C5:C6)"
        ws["D10"] = "=SUM(D5:D6)"
        ws["E10"] = "=SUM(E5:E6)"
        ws["F10"] = "=SUM(F5:F6)"

        ws["A11"] = "Subtotal Demais Regiões"
        ws["C11"] = "=SUM(C7:C8)"
        ws["D11"] = "=SUM(D7:D8)"
        ws["E11"] = "=SUM(E7:E8)"
        ws["F11"] = "=SUM(F7:F8)"

        ws["A12"] = "TOTAL GERAL"
        ws["C12"] = "=C10+C11"
        ws["D12"] = "=D10+D11"
        ws["E12"] = "=E10+E11"
        ws["F12"] = "=F10+F11"

        for r in [10, 11, 12]:
            ws.cell(row=r, column=3).number_format = "#,##0"
            ws.cell(row=r, column=4).number_format = "#,##0.00"
            ws.cell(row=r, column=5).number_format = "R$ #,##0.00"
            ws.cell(row=r, column=6).number_format = "#,##0.00"
            for c in range(1, 7):
                cell = ws.cell(row=r, column=c)
                cell.font = bold_font
                if r == 12:
                    cell.border = total_border
                    cell.fill = PatternFill(start_color="E1EBF5", end_color="E1EBF5", fill_type="solid")
                    
        ws["A14"] = "Metodologia de Cálculo e Parâmetros de Referência (TFD / SIGTAP SUS):"
        ws["A14"].font = bold_font
        
        ws["A15"] = f"• Parâmetros de cálculo de emissoes: {self.config['interior_co2']} kg de CO₂/km aplicado de forma global sobre os trajetos rodados."
        ws["A15"].font = italic_font
        
        ws["A16"] = f"• Deslocamento Terrestre (Interior): R$ {self.config['interior_km']:.2f} por km rodado (calculado sobre o trajeto total de ida e volta)."
        ws["A16"].font = italic_font
        
        ws["A17"] = f"• Ajuda de Custo para Alimentação (Benefício): R$ {self.config['interior_fixo']:.2f} fixos por atendimento (refeição sem pernoite)."
        ws["A17"].font = italic_font

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 16)

        wb.save(writer_path)
        self.logger.info("Aba de resumo adicionada e integrada com sucesso.")

    def run(self, parametro=None):    
        try:
            df_original = self.encontrar_e_ler_arquivo()
            df_processado = self.calcular_economicidade(df_original)
            
            caminho_saida = "output/economicidade-telessaude.xlsx"
            
            with pd.ExcelWriter(caminho_saida, engine='openpyxl') as writer:
                df_processado.to_excel(writer, sheet_name="Dados Calculados", index=False)
            
            self.criar_aba_resumo(caminho_saida, df_processado)
            self.logger.info(f"Processamento concluído com sucesso! Salvo em: {caminho_saida}")
            
        except Exception as e:
            self.logger.error(f"Falha durante a execução do processador: {e}")